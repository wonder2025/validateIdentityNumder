
from PyQt5 import QtCore, QtGui, QtWidgets

from validateUtil import validateUtil
import openpyxl
import os
from PyQt5.QtWidgets import QApplication, QAction, QFileDialog,  QTextEdit
class Ui_validateIdentityUI(QtWidgets.QMainWindow):
    def setupUi(self, validateIdentityUI):
        validateIdentityUI.setObjectName("validateIdentityUI")
        validateIdentityUI.resize(635, 448)
        self.centralwidget = QtWidgets.QWidget(validateIdentityUI)
        self.centralwidget.setObjectName("centralwidget")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(410, 280, 120, 80))
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.line = QtWidgets.QFrame(self.centralwidget)
        self.line.setGeometry(QtCore.QRect(60, 100, 501, 16))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(160, 70, 302, 25))
        self.widget.setObjectName("widget")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.widget)
        self.horizontalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.pushButton = QtWidgets.QPushButton(self.widget)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout_2.addWidget(self.pushButton)
        spacerItem = QtWidgets.QSpacerItem(138, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem)
        self.pushButton_2 = QtWidgets.QPushButton(self.widget)
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout_2.addWidget(self.pushButton_2)
        self.widget1 = QtWidgets.QWidget(self.centralwidget)
        self.widget1.setGeometry(QtCore.QRect(81, 30, 471, 25))
        self.widget1.setObjectName("widget1")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.widget1)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label = QtWidgets.QLabel(self.widget1)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.lineEdit = QtWidgets.QLineEdit(self.widget1)
        self.lineEdit.setObjectName("lineEdit")
        self.horizontalLayout.addWidget(self.lineEdit)
        self.pushButton_3 = QtWidgets.QPushButton(self.widget1)
        self.pushButton_3.setObjectName("pushButton_3")
        self.horizontalLayout.addWidget(self.pushButton_3)
        self.widget2 = QtWidgets.QWidget(self.centralwidget)
        self.widget2.setGeometry(QtCore.QRect(50, 130, 520, 231))
        self.widget2.setObjectName("widget2")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout(self.widget2)
        self.horizontalLayout_3.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.textBrowser = QtWidgets.QTextBrowser(self.widget2)
        self.textBrowser.setObjectName("textBrowser")
        self.horizontalLayout_3.addWidget(self.textBrowser)
        self.textBrowser_2 = QtWidgets.QTextBrowser(self.widget2)
        self.textBrowser_2.setObjectName("textBrowser_2")
        self.horizontalLayout_3.addWidget(self.textBrowser_2)
        validateIdentityUI.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(validateIdentityUI)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 635, 23))
        self.menubar.setObjectName("menubar")
        validateIdentityUI.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(validateIdentityUI)
        self.statusbar.setObjectName("statusbar")
        validateIdentityUI.setStatusBar(self.statusbar)

        self.retranslateUi(validateIdentityUI)
        self.pushButton_3.clicked.connect(self.open_file)
        QtCore.QMetaObject.connectSlotsByName(validateIdentityUI)

    def retranslateUi(self, validateIdentityUI):
        _translate = QtCore.QCoreApplication.translate
        validateIdentityUI.setWindowTitle(_translate("validateIdentityUI", "身份证校验工具V 0.0.1  IT开发工作室"))
        self.pushButton.setText(_translate("validateIdentityUI", "开始"))
        self.pushButton_2.setText(_translate("validateIdentityUI", "停止"))
        self.label.setText(_translate("validateIdentityUI", "文件路径"))
        self.pushButton_3.setText(_translate("validateIdentityUI", "打开"))
        self.textBrowser.setHtml(_translate("validateIdentityUI", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
"<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
"p, li { white-space: pre-wrap; }\n"
"</style></head><body style=\" font-family:\'SimSun\'; font-size:9pt; font-weight:400; font-style:normal;\">\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">使用说明：</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\"></span><span style=\" color:#ff0000;\">1.只支持后缀为.xlsx的excel文件的校验，默认第一行为标题，从第二行开始校验，只校验第一个sheet中的标题为‘身份证号码’的列；</span><span style=\" color:#5500ff;\"></span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">2.生成的文件位于原路径下，且校验结果位于生成校验结果EXCEL的最后几列；</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">3.身份证号校验规则为：</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">（1）长度必须为18位；</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">（2）前17位为数字，最后1位校验码为数字或X</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">（3）对前2位的省区划进行判断，属于31个省市自治区的区划之一；</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">（4）第7至14位为yyyymmdd出生日期，18&lt;=年龄&lt;=60；</span></p>\n"
"<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><span style=\" color:#5500ff;\">（5）最后一位校验码正确</span></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; color:#5500ff;\"><br /></p>\n"
"<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; color:#5500ff;\"><br /></p></body></html>"))



    def open_file(self):
        self.path = QFileDialog.getOpenFileName(self, 'Open file', './')
        self.lineEdit.setText( self.path[0])

    def setTextBrowser(self,text):
        self.textBrowser_2.append(text)

    def read_excel(self,begin_row, col, filePath, exportFilePath):
        try:
            self.setTextBrowser("正在载入excel...")
            rule = {
                1: "长度必须为18位",
                2: "前17位为数字，最后1位校验码为数字或X（如为x请转为X）",
                3: "前6位，可只对前2位的省区划进行判断，属于31个省市自治区的区划之一。（11,12,13,14,15，21,22,23,31,32,33,34,35,36,37,41,42,43,44,45,46,50,51,52,53,54,61,62,63,64,65)",
                4: "第7至14位为yyyymmdd出生日期，要求18<=年龄<=60，1<=mm<=12,1<=dd,if mm in (1/3/5/7/8/10/12) dd<=31 else if mm in (4/6/9/11) dd<=30 else if 闰年 dd<=28 else dd<=29",
                5: "1、将前面的身份证号码17位数分别乘以不同的系数。从第一位到第十七位的系数分别为：7－9－10－5－8－4－2－1－6－3－7－9－10－5－8－4－2。2、将这17位数字和系数相乘的结果相加。3、用加出来和除以11，余数为0－1－2－3－4－5－6－7－8－9－10这11个数字，其分别对应的最后一位身份证的号码为1－0－X －9－8－7－6－5－4－3－2。"
            }
            validatePass = "身份证校验通过"
            validateNotPass = "身份证校验不通过"
            col_name = "身份证号码"
            # 打开文件
            wb = openpyxl.load_workbook(filePath)
            sheetnames = wb.get_sheet_names()
            ws = wb.get_sheet_by_name(sheetnames[0])
            # 显示表名，表行数，表列数
            self.setTextBrowser("第一个sheet的名称为:" + ws.title)
            self.setTextBrowser("表规格:"+str(ws.max_row)+"行，"+str(ws.max_column)+"列")
            rows = ws.max_row
            cols = ws.max_column
            print("loaded excel")
            self.setTextBrowser("EXCEL加载完成，开始校验...")
            self.setTextBrowser("==========================")
            for i in range(begin_row, rows + 1):
                print("Validating No.%s" % (i))
                self.setTextBrowser("正在校验第"+str(i)+"行")
                value = ws.cell(row=i, column=col).value
                if len(value) != 18:
                    ws.cell(row=i, column=cols + 1).value = 0
                    ws.cell(row=i, column=cols + 1).value = validateNotPass
                    ws.cell(row=i, column=cols + 2).value = rule[1]
                    continue
                elif not (validateUtil.validate2(value)):  # 前17位为数字，最后1位校验码为数字或X（如为x请转为X）
                    ws.cell(row=i, column=cols + 1).value = validateNotPass
                    ws.cell(row=i, column=cols + 2).value = rule[2]
                    continue
                elif not (validateUtil.validate3(
                        value)):  # 前6位，可只对前2位的省区划进行判断，属于31个省市自治区的区划之一。（11,12,13,14,15，21,22,23,31,32,33,34,35,36,37,41,42,43,44,45,46,50,51,52,53,54,61,62,63,64,65)
                    ws.cell(row=i, column=cols + 1).value = validateNotPass
                    ws.cell(row=i, column=cols + 2).value = rule[3]
                    continue
                elif not (validateUtil.validate4(
                        value)):  # 第7至14位为yyyymmdd出生日期，要求18<=年龄<=60，1<=mm<=12,1<=dd,if mm in (1/3/5/7/8/10/12) dd<=31 else if mm in (4/6/9/11) dd<=30 else if 闰年 dd<=28 else dd<=29
                    ws.cell(row=i, column=cols + 1).value = 0
                    ws.cell(row=i, column=cols + 1).value = validateNotPass
                    ws.cell(row=i, column=cols + 2).value = rule[4]
                    continue
                elif not (validateUtil.validate5(
                        value)):  # 1、将前面的身份证号码17位数分别乘以不同的系数。从第一位到第十七位的系数分别为：7－9－10－5－8－4－2－1－6－3－7－9－10－5－8－4－2。2、将这17位数字和系数相乘的结果相加。3、用加出来和除以11，余数为0－1－2－3－4－5－6－7－8－9－10这11个数字，其分别对应的最后一位身份证的号码为1－0－X －9－8－7－6－5－4－3－2。
                    ws.cell(row=i, column=cols + 1).value = validateNotPass
                    ws.cell(row=i, column=cols + 2).value = rule[5]
                    continue
                else:
                    ws.cell(row=i, column=cols + 1).value = validatePass
            self.setTextBrowser("开始保存校验结果，请稍等...")
            if os.path.exists(exportFilePath):
                os.remove(exportFilePath)
            wb.save(exportFilePath)
            self.setTextBrowser("==========================")
            self.setTextBrowser("校验结束！")
        except Exception as err:
            print(err)
            self.setTextBrowser("==========================")
            self.setTextBrowser("校验失败!错误描述:"+err)