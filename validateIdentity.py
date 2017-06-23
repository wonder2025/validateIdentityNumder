import re
from validateUtil import validateUtil
import openpyxl
import os
def read_excel(sheetnum, begin_row,col,filePath,exportFilePath):
    print("loading excel")
    rule = {
        1:"长度必须为18位",
        2:"前17位为数字，最后1位校验码为数字或X（如为x请转为X）",
        3:"前6位，可只对前2位的省区划进行判断，属于31个省市自治区的区划之一。（11,12,13,14,15，21,22,23,31,32,33,34,35,36,37,41,42,43,44,45,46,50,51,52,53,54,61,62,63,64,65)",
        4:"第7至14位为yyyymmdd出生日期，要求18<=年龄<=60，1<=mm<=12,1<=dd,if mm in (1/3/5/7/8/10/12) dd<=31 else if mm in (4/6/9/11) dd<=30 else if 闰年 dd<=28 else dd<=29",
        5:"1、将前面的身份证号码17位数分别乘以不同的系数。从第一位到第十七位的系数分别为：7－9－10－5－8－4－2－1－6－3－7－9－10－5－8－4－2。2、将这17位数字和系数相乘的结果相加。3、用加出来和除以11，余数为0－1－2－3－4－5－6－7－8－9－10这11个数字，其分别对应的最后一位身份证的号码为1－0－X －9－8－7－6－5－4－3－2。"
    }
    validatePass= "身份证校验通过"
    validateNotPass= "身份证校验不通过"

    # 打开文件
    wb = openpyxl.load_workbook(filePath)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    # 显示表名，表行数，表列数
    print( "Work Sheet Titile:", ws.title)
    print("Work Sheet Rows:", ws.max_row)
    print("Work Sheet Cols:", ws.max_column)
    rows=ws.max_row
    cols=ws.max_column
    print("loaded excel")
    for i in range(begin_row, rows+1):
        print( "Validating No.%s" % (i))
        value =ws.cell(row=i, column=col).value
        if len(value)!=18:
            ws.cell(row=i, column=cols + 1).value=0
            ws.cell(row=i, column=cols + 1).value=validateNotPass
            ws.cell(row=i, column=cols + 2).value =rule[1]
            continue
        elif not(validateUtil.validate2(value)):#前17位为数字，最后1位校验码为数字或X（如为x请转为X）
            ws.cell(row=i, column=cols + 1).value=validateNotPass
            ws.cell(row=i, column=cols + 2).value =rule[2]
            continue
        elif not(validateUtil.validate3(value)):#前6位，可只对前2位的省区划进行判断，属于31个省市自治区的区划之一。（11,12,13,14,15，21,22,23,31,32,33,34,35,36,37,41,42,43,44,45,46,50,51,52,53,54,61,62,63,64,65)
            ws.cell(row=i, column=cols + 1).value=validateNotPass
            ws.cell(row=i, column=cols + 2).value =rule[3]
            continue
        elif not(validateUtil.validate4(value)):#第7至14位为yyyymmdd出生日期，要求18<=年龄<=60，1<=mm<=12,1<=dd,if mm in (1/3/5/7/8/10/12) dd<=31 else if mm in (4/6/9/11) dd<=30 else if 闰年 dd<=28 else dd<=29
            ws.cell(row=i, column=cols + 1).value=0
            ws.cell(row=i, column=cols + 1).value=validateNotPass
            ws.cell(row=i, column=cols + 2).value =rule[4]
            continue
        elif not(validateUtil.validate5(value)):#1、将前面的身份证号码17位数分别乘以不同的系数。从第一位到第十七位的系数分别为：7－9－10－5－8－4－2－1－6－3－7－9－10－5－8－4－2。2、将这17位数字和系数相乘的结果相加。3、用加出来和除以11，余数为0－1－2－3－4－5－6－7－8－9－10这11个数字，其分别对应的最后一位身份证的号码为1－0－X －9－8－7－6－5－4－3－2。
            ws.cell(row=i, column=cols + 1).value=validateNotPass
            ws.cell(row=i, column=cols + 2).value =rule[5]
            continue
        else:
            ws.cell(row=i, column=cols + 1).value =validatePass
    print("saving excel")
    if os.path.exists(exportFilePath):
        os.remove(exportFilePath)
    wb.save(exportFilePath)






if __name__ == '__main__':
    # filePath="D:\\身份证校验测试.xlsx"
    filePath="D:\\项目\身份校验\\身份证校验\\SICHUAN_20170619.xlsx"
    exportFilePath = "D:\\身份证校验结果.xlsx"
    read_excel(1, 2, 5,filePath,exportFilePath)




